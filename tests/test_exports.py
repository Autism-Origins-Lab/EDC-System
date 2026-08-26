from openpyxl import load_workbook

from app.database.queries.exports import EXPORT_SHEETS, export_patient_workbook


def test_export_patient_workbook_creates_expected_sheets(temp_database, tmp_path):
    output_path = tmp_path / "edc_export.xlsx"
    export_patient_workbook(output_path)

    workbook = load_workbook(output_path)
    assert set(EXPORT_SHEETS).issubset(set(workbook.sheetnames))
